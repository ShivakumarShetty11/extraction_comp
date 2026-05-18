import io
import json
import re
import time
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import anthropic
from anthropic import RateLimitError

ANTHROPIC_MODEL = "claude-haiku-4-5-20251001"

# ═══════════════════════════════════════════════════════════════════════════════
# Agent tool definitions — Anthropic format (input_schema, not parameters)
# ═══════════════════════════════════════════════════════════════════════════════

AGENT_TOOLS = [
    {
        "name": "scan_table_rows",
        "description": (
            "Scan sample rows from the table body to classify each row. "
            "Identifies: header rows (mostly text labels), data rows (numeric values), "
            "and column-number rows to skip (pattern like '(1) (2) (3)'). "
            "Call this first to understand the table structure."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "sample_rows": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "Array of rows; each row is an array of cell values (null for empty cells).",
                }
            },
            "required": ["sample_rows"],
        },
    },
    {
        "name": "detect_merged_cell_groups",
        "description": (
            "Detect merged-cell groups in a row by finding runs of consecutive identical values. "
            "In Excel, merged cells are filled with the same value — so 'AGE IN YEARS' appearing "
            "12 times in a row means that header spans 12 columns. "
            "Call this for each header row to understand multi-level column structure."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "row": {
                    "type": "array",
                    "description": "A single table row as an array of cell values.",
                },
                "row_index": {
                    "type": "integer",
                    "description": "0-based index of this row in the sample.",
                },
            },
            "required": ["row", "row_index"],
        },
    },
    {
        "name": "build_flat_column_names",
        "description": (
            "Build flat column names by combining multi-level header rows and merge group info. "
            "For a group header like 'AGE IN YEARS' spanning cols 4-15, and sub-columns "
            "'<1','1-4','5-14'... produces 'AGE IN YEARS_<1', 'AGE IN YEARS_1-4', etc. "
            "Call this last after you know which rows are headers and their merge patterns."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "header_rows": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": "The header rows (arrays of cell values).",
                },
                "merge_groups": {
                    "type": "array",
                    "description": "List of merge-group dicts from detect_merged_cell_groups calls.",
                },
                "n_cols": {
                    "type": "integer",
                    "description": "Exact number of columns required in output.",
                },
            },
            "required": ["header_rows", "merge_groups", "n_cols"],
        },
    },
]


# ═══════════════════════════════════════════════════════════════════════════════
# Tool executor functions (deterministic, no LLM)
# ═══════════════════════════════════════════════════════════════════════════════

def _exec_scan_table_rows(sample_rows: List[List[Any]]) -> str:
    results = []
    for i, row in enumerate(sample_rows):
        non_none = [v for v in row if v is not None]
        if not non_none:
            results.append({"row": i, "classification": "blank"})
            continue
        texts = sum(1 for v in non_none if isinstance(v, str))
        nums = sum(1 for v in non_none if isinstance(v, (int, float)))
        empty = len(row) - len(non_none)
        col_num = sum(1 for v in non_none if re.match(r"^\(\d+\)$", str(v).strip()))
        is_col_num_row = col_num > len(non_none) * 0.5
        results.append({
            "row": i,
            "classification": "col_number_row" if is_col_num_row
                              else ("header_candidate" if texts > nums else "data_row"),
            "text_cells": texts,
            "number_cells": nums,
            "empty_cells": empty,
            "sample_values": [str(v)[:25] for v in non_none[:5]],
        })
    return json.dumps(results, indent=2)


def _exec_detect_merged_cell_groups(row: List[Any], row_index: int) -> str:
    groups = []
    i = 0
    while i < len(row):
        v = row[i]
        if v is None:
            i += 1
            continue
        j = i + 1
        while j < len(row) and row[j] == v:
            j += 1
        groups.append({
            "value": str(v)[:50],
            "start_col": i,
            "end_col": j - 1,
            "span": j - i,
            "is_merged": j - i > 1,
        })
        i = j
    return json.dumps({"row": row_index, "groups": groups}, indent=2)


def _exec_build_flat_column_names(
    header_rows: List[List[Any]], merge_groups: List[Dict], n_cols: int
) -> str:
    if not header_rows:
        return json.dumps([f"Col_{i+1}" for i in range(n_cols)])

    if len(header_rows) == 1:
        cols = [str(v).strip() if v is not None else f"Col_{i+1}" for i, v in enumerate(header_rows[0])]
        while len(cols) < n_cols:
            cols.append(f"Col_{len(cols)+1}")
        return json.dumps(cols[:n_cols])

    row0 = list(header_rows[0])
    row1 = list(header_rows[1]) if len(header_rows) > 1 else []

    # Build a col→group_label map from merge_groups for row 0.
    # merge_groups is a list of dicts returned by detect_merged_cell_groups;
    # each dict has {"row": int, "groups": [{value, start_col, end_col, span, is_merged}]}.
    group_label: Dict[int, str] = {}
    for mg in merge_groups:
        if not isinstance(mg, dict):
            continue
        if mg.get("row", -1) != 0:
            continue
        for g in mg.get("groups", []):
            if g.get("is_merged"):
                for col_idx in range(g["start_col"], g["end_col"] + 1):
                    group_label[col_idx] = str(g["value"]).strip()

    cols = []
    for i in range(n_cols):
        v0 = row0[i] if i < len(row0) else None
        v1 = row1[i] if i < len(row1) else None

        # Prefer the authoritative group label from merge_groups for row 0;
        # fall back to counting repeated values if merge_groups wasn't provided.
        group = group_label.get(i)
        if group is None and v0 is not None:
            same_count = sum(1 for x in row0 if x == v0)
            if same_count > 1:
                group = str(v0).strip()

        if group and v1 is not None:
            cols.append(f"{group}_{str(v1).strip()}")
        elif v1 is not None:
            cols.append(str(v1).strip())
        elif v0 is not None:
            cols.append(str(v0).strip())
        else:
            cols.append(f"Col_{i+1}")

    return json.dumps(cols)


def _parse_retry_delay(error: RateLimitError) -> Optional[float]:
    """
    Extract the server-suggested retry delay from a 429 error body.
    Gemini embeds a retryDelay field (e.g. "24s") in the error details.
    Returns seconds as a float, or None if not found.
    """
    try:
        body = getattr(error, "body", None) or {}
        for detail in body.get("error", {}).get("details", []):
            raw = detail.get("retryDelay", "")
            if raw:
                m = re.match(r"([\d.]+)", str(raw))
                if m:
                    return float(m.group(1)) + 3  # add 3 s buffer
    except Exception:
        pass
    return None


def _is_daily_quota_error(error: RateLimitError) -> bool:
    """
    Return True when the quota that's exhausted resets daily (not per-minute).
    Retrying in the same session won't help — fail fast so the caller falls
    back to heuristic immediately instead of burning 60+ seconds waiting.
    """
    try:
        body = getattr(error, "body", None) or {}
        for detail in body.get("error", {}).get("details", []):
            for v in detail.get("violations", []):
                if "PerDay" in v.get("quotaId", ""):
                    return True
        # Also check the plain message string as a fallback
        msg = body.get("error", {}).get("message", "")
        if "PerDay" in msg or "per_day" in msg.lower():
            return True
    except Exception:
        pass
    return False


def _call_with_retry(client: anthropic.Anthropic, max_retries: int = 3, **kwargs):
    """
    Call client.messages.create with smart backoff on RateLimitError.
    - Checks Retry-After header for the suggested wait time.
    - Falls back to 30 s / 60 s exponential backoff when no header is present.
    """
    for attempt in range(max_retries):
        try:
            return client.messages.create(**kwargs)
        except RateLimitError as e:
            if attempt == max_retries - 1:
                raise
            wait = _parse_retry_delay(e) or (30 * (2 ** attempt))
            print(f"Rate limit — waiting {wait:.0f}s before retry {attempt + 2}/{max_retries} …")
            time.sleep(wait)


def _extract_json_with_key(text: str, required_key: str) -> Optional[Dict]:
    """Find the first valid JSON object in `text` that contains `required_key`."""
    # Strip markdown fences first
    text = re.sub(r"```[a-z]*\n?", "", text).strip().rstrip("`")
    decoder = json.JSONDecoder()
    idx = 0
    while idx < len(text):
        # Find the next opening brace
        start = text.find("{", idx)
        if start == -1:
            break
        try:
            obj, end_pos = decoder.raw_decode(text, start)
            if isinstance(obj, dict) and required_key in obj:
                return obj
            idx = start + 1
        except json.JSONDecodeError:
            idx = start + 1
    return None


TOOL_MAP = {
    "scan_table_rows": lambda a: _exec_scan_table_rows(a["sample_rows"]),
    "detect_merged_cell_groups": lambda a: _exec_detect_merged_cell_groups(a["row"], a["row_index"]),
    "build_flat_column_names": lambda a: _exec_build_flat_column_names(
        a["header_rows"], a.get("merge_groups", []), a["n_cols"]
    ),
}


# ═══════════════════════════════════════════════════════════════════════════════
# TableExtractor — shared openpyxl logic + two structure-analysis strategies
# ═══════════════════════════════════════════════════════════════════════════════

class TableExtractor:
    def __init__(self, api_key: Optional[str] = None, use_agent: bool = False):
        self.client = anthropic.Anthropic(api_key=api_key)
        self.use_agent = use_agent

    # ── Public API ────────────────────────────────────────────────────────────

    def extract_from_file(self, file_content: bytes, filename: str) -> List[Dict]:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception as e:
            raise ValueError(f"Cannot open Excel file: {e}")

        all_tables: List[Dict] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_tables.extend(self._process_sheet(ws, sheet_name, filename))
        return all_tables

    # ── Sheet processing ──────────────────────────────────────────────────────

    def _process_sheet(self, ws, sheet_name: str, filename: str) -> List[Dict]:
        grid = self._filled_grid(ws)
        if not grid:
            return []
        results = []
        blocks = self._find_blocks(grid)
        for idx, (start, end) in enumerate(blocks):
            try:
                tbl = self._extract_table(grid, start, end, sheet_name, filename, idx)
                if tbl and tbl["row_count"] > 0:
                    results.append(tbl)
            except Exception as e:
                print(f"[{sheet_name}] block {start}-{end} error: {e}")
            # Pause between tables so back-to-back LLM calls don't hit per-minute limits.
            # Skip the pause after the last table in this sheet.
            if idx < len(blocks) - 1:
                time.sleep(3)
        return results

    # ── Grid builder ──────────────────────────────────────────────────────────

    def _filled_grid(self, ws) -> List[List[Any]]:
        merge_lookup: Dict[Tuple[int, int], Any] = {}
        for mrng in ws.merged_cells.ranges:
            master = ws.cell(mrng.min_row, mrng.min_col).value
            for r in range(mrng.min_row, mrng.max_row + 1):
                for c in range(mrng.min_col, mrng.max_col + 1):
                    if r != mrng.min_row or c != mrng.min_col:
                        merge_lookup[(r, c)] = master

        max_row = max_col = 0
        for row in ws.iter_rows():
            for cell in row:
                v = merge_lookup.get((cell.row, cell.column), cell.value)
                if v is not None and str(v).strip():
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)

        if not max_row:
            return []
        return [
            [merge_lookup.get((r, c), ws.cell(r, c).value) for c in range(1, max_col + 1)]
            for r in range(1, max_row + 1)
        ]

    # ── Block detection ───────────────────────────────────────────────────────

    @staticmethod
    def _blank(row: List[Any]) -> bool:
        return all(v is None or str(v).strip() == "" for v in row)

    @staticmethod
    def _has_table_marker(row: List[Any]) -> bool:
        text = " ".join(str(v) for v in row if v is not None)
        return bool(re.search(r"\bTABLE\s*[:\-]?\s*[A-Z\d]", text, re.IGNORECASE))

    def _find_blocks(self, grid: List[List[Any]]) -> List[Tuple[int, int]]:
        markers = [i for i, r in enumerate(grid) if self._has_table_marker(r)]
        if markers:
            blocks = []
            for j, start in enumerate(markers):
                limit = markers[j + 1] if j + 1 < len(markers) else len(grid)
                end = limit - 1
                while end > start and self._blank(grid[end]):
                    end -= 1
                if end > start:
                    blocks.append((start, end))
            return blocks

        blocks, current, blanks = [], None, 0
        for i, row in enumerate(grid):
            if self._blank(row):
                blanks += 1
                if blanks >= 2 and current is not None:
                    end = i - blanks
                    if end > current:
                        blocks.append((current, end))
                    current, blanks = None, 0
            else:
                blanks = 0
                if current is None:
                    current = i
        if current is not None:
            end = len(grid) - 1
            while end > current and self._blank(grid[end]):
                end -= 1
            if end > current:
                blocks.append((current, end))
        return blocks

    # ── Table extraction ──────────────────────────────────────────────────────

    def _extract_table(
        self, grid, start: int, end: int, sheet_name: str, filename: str, idx: int
    ) -> Optional[Dict]:
        block = grid[start : end + 1]
        title, description, body_start = self._strip_title_desc(block)
        body = block[body_start:]
        if not body:
            return None

        n_cols = max(len(r) for r in body)

        agent_steps: Optional[List[Dict]] = None
        try:
            if self.use_agent:
                structure, agent_steps = self._agent_structure(body, title, description, n_cols)
            else:
                structure = self._direct_llm_structure(body, title, description, n_cols)
        except Exception as e:
            print(f"Structure analysis failed ({e}), using heuristic")
            structure = self._heuristic_structure(body, n_cols)
            if self.use_agent:
                # Always surface a steps list so the UI shows what happened
                agent_steps = [{"type": "fallback", "content": f"Agent error: {str(e)[:400]}. Fell back to heuristic."}]

        header_rows: int = structure.get("header_rows", 1)
        skip_set: set = set(structure.get("skip_rows", []))
        columns: List[str] = list(structure.get("columns", []))

        while len(columns) < n_cols:
            columns.append(f"Col_{len(columns)+1}")
        columns = columns[:n_cols]

        # Deduplicate column names
        seen: Dict[str, int] = {}
        deduped = []
        for col in columns:
            if col in seen:
                seen[col] += 1
                deduped.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 0
                deduped.append(col)
        columns = deduped

        rows = []
        for i, row in enumerate(body[header_rows:], start=header_rows):
            if i in skip_set or self._blank(row):
                continue
            padded = list(row) + [None] * (n_cols - len(row))
            row_dict = {}
            for j, col in enumerate(columns):
                v = padded[j]
                if hasattr(v, "item"):
                    v = v.item()
                row_dict[col] = v
            rows.append(row_dict)

        def _ser(row):
            out = []
            for v in row:
                if v is None:
                    out.append(None)
                elif hasattr(v, "item"):
                    out.append(v.item())
                else:
                    out.append(str(v).strip())
            return out

        # Raw header rows before AI flattening (multi-level structure)
        raw_header_rows = [_ser(row) for row in body[:header_rows]]

        # Column-index rows like (1)(2)(3) that were skipped
        raw_col_num_rows = [_ser(body[i]) for i in sorted(skip_set) if i < len(body)]

        # Footer notes: text-heavy rows at the end of the body after data
        raw_notes: List[str] = []
        for row in reversed(body[header_rows:]):
            non_none = [v for v in row if v is not None and str(v).strip()]
            if not non_none:
                continue
            nums = sum(1 for v in non_none if isinstance(v, (int, float)))
            if nums > len(non_none) * 0.3:
                break  # hit a data row — stop scanning
            text = " ".join(str(v).strip() for v in non_none)
            if len(text) > 8:
                raw_notes.insert(0, text)
            if len(raw_notes) >= 10:
                break

        result: Dict = {
            "id": f"{filename}__{sheet_name}__{idx}",
            "title": title or f"Table {idx+1}",
            "description": description,
            "sheet": sheet_name,
            "filename": filename,
            "columns": columns,
            "rows": rows,
            "row_count": len(rows),
            "raw_header_rows": raw_header_rows,
            "raw_col_num_rows": raw_col_num_rows,
            "raw_notes": raw_notes,
        }
        if agent_steps is not None:
            result["agent_steps"] = agent_steps
        return result

    # ── Title / description extraction ────────────────────────────────────────

    def _strip_title_desc(self, block: List[List[Any]]) -> Tuple[str, str, int]:
        def row_text(row):
            seen_vals, parts = set(), []
            for v in row:
                if v is None:
                    continue
                sv = str(v).strip()
                if sv and sv not in seen_vals:
                    seen_vals.add(sv)
                    parts.append(sv)
            return " ".join(parts)

        non_blank = [(i, r) for i, r in enumerate(block) if not self._blank(r)]
        title = description = ""
        body_start = 0
        if not non_blank:
            return title, description, body_start

        i0, r0 = non_blank[0]
        title = row_text(r0)
        body_start = i0 + 1

        if len(non_blank) > 1:
            i1, r1 = non_blank[1]
            text1 = row_text(r1)
            non_none = [v for v in r1 if v is not None]
            n_nums = sum(1 for v in non_none if isinstance(v, (int, float)))
            if text1 and len(text1) > 10 and n_nums < len(non_none) / 2:
                description = text1
                body_start = i1 + 1

        while body_start < len(block) and self._blank(block[body_start]):
            body_start += 1
        return title, description, body_start

    # ─────────────────────────────────────────────────────────────────────────
    # STRATEGY A — Direct LLM (single prompt → single response)
    # ─────────────────────────────────────────────────────────────────────────

    def _direct_llm_structure(
        self, body: List[List[Any]], title: str, description: str, n_cols: int
    ) -> Dict:
        n_sample = min(10, len(body))
        lines = []
        for i, row in enumerate(body[:n_sample]):
            vals = "\t".join("" if v is None else str(v).strip() for v in row)
            lines.append(f"Row {i+1}:\t{vals}")
        grid_text = "\n".join(lines)

        prompt = f"""Analyze this government statistical table body and return its structure.

Table: {title}
{f"Description: {description}" if description else ""}
Columns: {n_cols}, Total rows: {len(body)}

First {n_sample} rows (tab-separated):
{grid_text}

Return JSON with:
- "header_rows": count of rows at start that are column headers (not data)
- "skip_rows": 0-based row indices to skip (e.g. column-number rows like "(1)(2)(3)")
- "columns": exactly {n_cols} flat column names

For multi-level headers (group row + sub-column row), combine: "AGE_<1", "AGE_1-4"
Return ONLY valid JSON, no markdown fences."""

        resp = _call_with_retry(
            self.client,
            model=ANTHROPIC_MODEL,
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        text = re.sub(r"```[a-z]*\n?", "", text).strip().rstrip("`").strip()
        return json.loads(text)

    # ─────────────────────────────────────────────────────────────────────────
    # STRATEGY B — AI Agent (ReAct loop with tool calling)
    # ─────────────────────────────────────────────────────────────────────────

    def _agent_structure(
        self, body: List[List[Any]], title: str, description: str, n_cols: int
    ) -> Tuple[Dict, List[Dict]]:
        """Returns (structure_dict, agent_steps_list)."""

        n_sample = min(6, len(body))   # keep token count low on free-tier APIs
        sample = [
            [v if isinstance(v, (int, float, str, type(None))) else str(v) for v in row]
            for row in body[:n_sample]
        ]

        system_msg = (
            "You are an expert data analyst extracting table structure from government Excel reports. "
            "You have three tools available:\n"
            "1. scan_table_rows — call this FIRST to classify each row (header vs data vs skip)\n"
            "2. detect_merged_cell_groups — call this for each header row to find merged column groups\n"
            "3. build_flat_column_names — call this LAST with all header rows and merge info\n\n"
            "After using the tools, return your final answer as a plain JSON object (no markdown):\n"
            '{"header_rows": <int>, "skip_rows": [<int>,...], "columns": [<str>,...]}\n'
            f"The columns array MUST contain exactly {n_cols} entries."
        )

        user_msg = (
            f"Table: {title}\n"
            f"{f'Description: {description}' if description else ''}\n"
            f"Total body rows: {len(body)}, Columns: {n_cols}\n\n"
            f"Sample rows (JSON):\n{json.dumps(sample, indent=2)}\n\n"
            f"Use your tools step-by-step, then return the final JSON structure."
        )

        # Anthropic: system is a top-level param; messages start with user turn
        messages = [{"role": "user", "content": user_msg}]

        steps: List[Dict] = []
        structure: Optional[Dict] = None
        tools_called: set = set()
        required_tools = {"scan_table_rows", "detect_merged_cell_groups", "build_flat_column_names"}
        MAX_ITER = 12

        for iteration in range(MAX_ITER):
            all_tools_done = required_tools.issubset(tools_called)
            # Anthropic tool_choice: {"type":"any"} forces at least one tool call
            tool_choice = {"type": "auto"} if all_tools_done else {"type": "any"}

            resp = _call_with_retry(
                self.client,
                model=ANTHROPIC_MODEL,
                system=system_msg,
                messages=messages,
                tools=AGENT_TOOLS,
                tool_choice=tool_choice,
                max_tokens=2048,
            )

            # Anthropic returns a list of content blocks (TextBlock / ToolUseBlock)
            text_parts = [b.text for b in resp.content if hasattr(b, "text") and b.text]
            tool_uses  = [b for b in resp.content if b.type == "tool_use"]
            text = " ".join(text_parts)

            # Record thought step
            thought_step: Dict = {"type": "thought", "content": text}
            if tool_uses:
                thought_step["tool_calls"] = [
                    {"tool": tu.name, "input_raw": json.dumps(tu.input)}
                    for tu in tool_uses
                ]
            steps.append(thought_step)

            # Append assistant turn (full content block list keeps context intact)
            messages.append({"role": "assistant", "content": resp.content})

            # No tool uses → final answer turn
            if not tool_uses:
                parsed = _extract_json_with_key(text, "header_rows")
                if parsed is not None:
                    structure = parsed
                    steps.append({"type": "final_answer", "content": structure})
                    break
                messages.append({
                    "role": "user",
                    "content": (
                        "Return ONLY a JSON object — no prose, no markdown — with exactly these keys: "
                        f"header_rows (int), skip_rows (list of ints), columns (list of exactly {n_cols} strings)."
                    ),
                })
                continue

            # Execute each tool use, collect all results into a single user message
            tool_result_blocks = []
            for tu in tool_uses:
                tools_called.add(tu.name)
                try:
                    output = TOOL_MAP[tu.name](tu.input)
                except Exception as e:
                    output = json.dumps({"error": str(e)})

                steps.append({
                    "type": "tool_result",
                    "tool": tu.name,
                    "input": tu.input,
                    "output": output,
                })
                tool_result_blocks.append({
                    "type": "tool_result",
                    "tool_use_id": tu.id,
                    "content": output,
                })

            # Anthropic requires all tool results in one user turn
            messages.append({"role": "user", "content": tool_result_blocks})

            # Brief pause between iterations to respect per-minute rate limits
            time.sleep(2)

        if structure is None:
            structure = self._heuristic_structure(body, n_cols)
            steps.append({"type": "fallback", "content": "Agent did not return valid JSON after all iterations. Used heuristic fallback."})

        return structure, steps

    # ── LLM-based category metadata extraction ───────────────────────────────

    def extract_category_metadata(
        self,
        title: str,
        description: str,
        raw_header_rows: List[List],
        columns: List[str],
        sample_rows: List[Dict],
        raw_notes: List[str],
    ) -> List[Dict]:
        """
        Ask the LLM to semantically identify all categorical dimensions present in
        the table — e.g. Area Type (Rural/Urban), Care Setting (Institution/Domiciliary),
        Gender (Male/Female), Geography (State, District), Age Group, etc.
        Returns a list of category dicts: {name, description, values:[{value,code,description}]}
        """
        # Format original header matrix so the LLM sees the raw multi-level structure
        header_lines = []
        for i, row in enumerate(raw_header_rows):
            deduped: List[str] = []
            seen_in_row: set = set()
            for v in row:
                sv = str(v).strip() if v is not None else ""
                if sv and sv not in seen_in_row:
                    deduped.append(sv)
                    seen_in_row.add(sv)
            if deduped:
                header_lines.append(f"  Header level {i+1}: {' | '.join(deduped)}")

        # Sample the first few data rows so the LLM can see row-dimension values
        data_sample_lines = []
        for row in sample_rows[:6]:
            vals = " | ".join(str(v) for v in list(row.values())[:6] if v is not None)
            if vals:
                data_sample_lines.append(f"  {vals}")

        prompt = f"""You are analyzing a government statistical table to build a metadata catalogue.

Table: {title}
{f'Description: {description}' if description else ''}

Original multi-level column headers (before flattening):
{chr(10).join(header_lines) if header_lines else '  (single-level headers)'}

Flat column names after AI extraction ({len(columns)} total):
  {', '.join(columns[:30])}{'…' if len(columns) > 30 else ''}

First few data rows (first 6 values each):
{chr(10).join(data_sample_lines) if data_sample_lines else '  (no sample)'}
{f"Source notes: {'; '.join(raw_notes[:4])}" if raw_notes else ''}

Your task: identify EVERY categorical dimension in this table.
A "category" is any grouping variable — examples:
- Column-group headers like RURAL / URBAN (spans multiple sub-columns)
- Sub-group headers like INSTITUTION / DOMICILIARY
- Cross-tabulation variables like Male / Female, Age bands, Year
- Row-dimension labels like State, District, Block, Taluk

Rules:
- Each category name must be UNIQUE — do not repeat the same category twice
- Each value within a category must be UNIQUE — list each value only once
- Merge overlapping categories into one (e.g. "Area" and "Location Type" are the same)
- Omit "Total" or "Grand Total" rows — those are aggregates, not categories

For each category:
- Give it a clear human-readable name (e.g. "Area Type", "Care Setting", "Gender")
- Write a one-sentence description of what it classifies
- List every unique value you can identify from the headers and data sample
- Generate a short UPPERCASE_UNDERSCORE code for each value
- Write a brief plain-English description for each value

Return ONLY a valid JSON object — no prose, no markdown fences:
{{
  "categories": [
    {{
      "name": "Area Type",
      "description": "Classifies data by geographic area type",
      "values": [
        {{"value": "Rural", "code": "RURAL", "description": "Data from rural areas outside municipal limits"}},
        {{"value": "Urban", "code": "URBAN", "description": "Data from urban areas within municipal limits"}}
      ]
    }}
  ]
}}"""

        resp = _call_with_retry(
            self.client,
            model=ANTHROPIC_MODEL,
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        text = resp.content[0].text.strip()
        parsed = _extract_json_with_key(text, "categories")
        if parsed is None:
            return []
        return parsed.get("categories", [])

    # ── Heuristic fallback ────────────────────────────────────────────────────

    @staticmethod
    def _heuristic_structure(body: List[List[Any]], n_cols: int) -> Dict:
        if not body:
            return {"header_rows": 0, "skip_rows": [], "columns": [f"Col_{i+1}" for i in range(n_cols)]}
        skip_rows = []
        for i, row in enumerate(body[:6]):
            non_none = [v for v in row if v is not None]
            if not non_none:
                continue
            col_nums = sum(1 for v in non_none if re.match(r"^\(\d+\)$", str(v).strip()))
            if col_nums > len(non_none) * 0.5:
                skip_rows.append(i)
        header_rows = 1
        for i, row in enumerate(body[:5]):
            non_none = [v for v in row if v is not None]
            nums = sum(1 for v in non_none if isinstance(v, (int, float)))
            if nums > len(non_none) * 0.4 and i > 0:
                header_rows = i
                break
        columns = [str(v).strip() if v is not None else f"Col_{i+1}" for i, v in enumerate(body[0])]
        while len(columns) < n_cols:
            columns.append(f"Col_{len(columns)+1}")
        return {"header_rows": header_rows, "skip_rows": skip_rows, "columns": columns}
